from tkinter import *
from tkinter import filedialog
from tkinter import ttk
from openpyxl import Workbook, load_workbook

from models.student_model import Student

class AttendanceKeeper(Frame):
    def __init__(self, parent):
        self.users = []
        self.sections = []

        self.section_students = []
        self.attended_students = []

        self.selected_students = []

        Frame.__init__(self, parent)
        self.initUI(parent)
    def exportAttendedStudents(self):
        file_type = self.file_type_box.get()
        section = self.section_box.get()
        entry = self.entry.get()
        if file_type == "txt":
            with open("{} Week {}.txt".format(section,entry), "w+",  encoding='utf8') as file:
                file.write("ID    Name    Departman\n")
                for student in self.attended_students:
                    file.write("{}    {}    {}\n".format(student.Id, student.Name, student.Department))
        elif file_type == "xls":
            workbook = Workbook()
            worksheet = workbook.active

            worksheet['A1'] = 'ID'
            worksheet['B1'] = 'Name'
            worksheet['C1'] = 'Department'

            for student in self.attended_students:
                worksheet.append([student.Id, student.Name, student.Department])

            workbook.save("{} Week {}.xls".format(section,entry))
        else:
           raise BaseException("File type is not supported!")
    def fillSectionStudentsBySection(self, event):
        self.section_students = []
        self.select_list.delete(0, END)
        self.attended_list.delete(0, END)
        for index, student in enumerate(self.users):
            if student.Section == self.section_box_value.get():
                self.section_students.append(student)
                self.select_list.insert(index, student.toString())
    def callbackRemoveSelectedStudents(self):
        selected_students = self.attended_list.curselection()
        for index in selected_students:
            self.attended_students.pop(index)
            self.attended_list.delete(index)

    def attendStudentsToList(self):
        for index, student in enumerate(self.selected_students):
            if student not in self.attended_students:
                self.attended_students.append(student)
                self.attended_list.insert(index, student.toString())

    def callbackSelectedItems(self, event):
        self.selected_students = []
        selected_students = self.select_list.curselection()
        for index in selected_students:
            self.selected_students.append(self.section_students[index])

    def buildUI(self):
        for index, student in enumerate(self.users):
            self.section_box['values'] = self.sections
        self.section_box.current(0)
        self.fillSectionStudentsBySection(None)
    
    def sortSection(self,eleman):
        return int(''.join(filter(str.isdigit, eleman))) 
        
    def readAllData(self,path):
        swb = load_workbook(path)["ENGR 102_studentList.Raw.3.3.20"]
        for index, row in enumerate(swb.iter_rows(values_only=True, min_row=2)):
            if row[3] not in self.sections:
                #Tüm sectionları listeye ekle
                self.sections.append(row[3])
            self.users.append(Student(row))


        self.sections.sort(key=self.sortSection) # listenin sonundaki sayıya göre sıralar.
        self.buildUI()

    def browseFiles(self):
        filename = filedialog.askopenfilename(initialdir="/", title="Select a File", filetypes=(("Excell files", ["*.xls","*.xlsx"]),))
        self.readAllData(filename)

    def initUI(self, parent):
        #entry
        self.entry = Entry(parent, width=9, font=('Arial', 14))
        #combobox
        self.section_box_value = StringVar()
        self.file_type_box_value = StringVar()
        
        #listbox
        self.select_list = Listbox(width=40, height=6 ,selectmode = "multiple")
        self.attended_list = Listbox(width=40, height=6)
        
        #labelleri bulunduğu kısım
        self.title = Label(text="AttendanceKeeper v1.0", font=("", 22, "bold"))
        self.select_file = Label(text="Select Student List Excell File:", font=("", 12, "bold"))
        self.select_std = Label(text="Select a Student:", font=("", 12, "bold"))
        self.section = Label(text="Section:", font=("", 12, "bold"))
        self.attended_std = Label(text="Attended Student:", font=("", 12, "bold"))
        self.file_type = Label(text="Please Select File Type:", font=("", 12, "bold"))
        self.week = Label(text="Please Enter Week:", font=("", 12, "bold"))
        
        #buttonların bulunduğu kısım
        self.file_btn = Button(text="Import List", command=self.browseFiles, width=19, height=1)
        self.add_btn = Button(text="Add =>",width=19, height=2, command=self.attendStudentsToList)
        self.remove_btn = Button(text="<= Remove",width=19, height=2, command=self.callbackRemoveSelectedStudents)
        self.export_file = Button(text="Export File",width=19, height=1, command=self.exportAttendedStudents)
        
        #Combobox
        self.section_box = ttk.Combobox(textvariable=self.section_box_value)
        self.file_type_box = ttk.Combobox(textvariable=self.file_type_box_value, width=5)
        self.file_type_box['values'] = ("txt", "csv", "xls")
        self.file_type_box.current(0)
        self.section_box.bind("<<ComboboxSelected>>", self.fillSectionStudentsBySection)
        self.select_list.bind("<<ListboxSelect>>", self.callbackSelectedItems)
        
        #ekran
        #row 0
        self.title.grid(row=0, column=0,columnspan=3)
        
        #row 1
        self.file_btn.grid(column=1, row=1)
        self.select_file.grid(column=0, row=1)
        
        #row 2
        self.select_std.grid(column=0, row=2, sticky=W)
        self.section.grid(column=1, row=2)
        self.attended_std.grid(column=2, row=2, sticky=W)
        
        #row 3
        self.section_box.grid(column=1, row=3)
        self.select_list.grid(column=0, row=3, rowspan=3)
        self.attended_list.grid(column=2, row=3, rowspan=3)
        
        #row 4
        self.add_btn.grid(column=1, row=4)
        
        #row 5
        self.remove_btn.grid(column=1, row=5)
        
        #row 6
        self.file_type.grid(column=0, row=6, sticky=W)
        self.file_type_box.grid(column=0, row=6, sticky=E)
        self.week.grid(column=1, row=6)
        self.entry.grid(column=2, row=6, sticky=W)
        self.export_file.grid(column=2, row=6, sticky=E)
        


def main():
    root = Tk()
    root.geometry("645x225+695+290")
    root.resizable(False, False)
    AttendanceKeeper(root)
    root.mainloop()
    
main()